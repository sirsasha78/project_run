from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework import status
from openpyxl import load_workbook

from artifacts.models import CollectibleItem
from artifacts.serializers import CollectibleItemSerializer


class CollectibleItemViewSet(viewsets.ReadOnlyModelViewSet):
    """Набор представлений для отображения коллекционных предметов.
    Предоставляет только операции чтения (просмотр одного или списка объектов)
    для модели CollectibleItem. Используется для безопасного доступа к данным
    без возможности их изменения через API.
    Атрибуты:
        queryset (QuerySet): Набор объектов модели CollectibleItem,
            доступных для просмотра.
        serializer_class (Serializer): Класс сериализатора, используемый
            для преобразования объектов модели в JSON и обратно."""

    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer


class UploadFileView(APIView):
    """Представление для загрузки и обработки Excel-файла с данными о коллекционных предметах.
    Принимает POST-запрос с файлом в формате .xlsx, считывает данные с активного листа,
    валидирует их с помощью сериализатора и сохраняет в базу данных. Невалидные строки
    собираются и возвращаются в ответе.
    Атрибуты:
        serializer_class (CollectibleItemSerializer): Сериализатор для валидации
            и сохранения данных каждого предмета."""

    serializer_class = CollectibleItemSerializer

    def post(self, request: Request, *args, **kwargs) -> Response:
        """Обрабатывает POST-запрос на загрузку файла.
        Ожидает файл в поле 'file' multipart-формы. Файл должен быть в формате .xlsx.
        Считывает данные со второго ряда (первый — заголовки) и пытается создать объекты
        модели по каждой строке. В случае ошибки валидации строка добавляется в список
        невалидных данных."""

        upload_file = request.FILES.get("file")
        if not upload_file:
            return Response(
                {"message": "Файл не найден"}, status=status.HTTP_400_BAD_REQUEST
            )

        wb = load_workbook(upload_file)
        page = wb.active
        invalid_data = []

        for row in page.iter_rows(min_row=2, values_only=True):
            data = {
                "name": row[0],
                "uid": row[1],
                "value": row[2],
                "latitude": row[3],
                "longitude": row[4],
                "picture": row[5],
            }
            serializer = self.serializer_class(data=data)
            if serializer.is_valid():
                serializer.save()
            else:
                invalid_data.append(list(row))
        return Response(invalid_data, status=status.HTTP_200_OK)
